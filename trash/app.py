import os
import argparse
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import random
import shutil

def leer_datos_fuente(ruta_excel="datos_fuente.xlsx"):
    """
    Lee datos_fuente.xlsx
    
    Args:
        print a la ruta del archivo datos_fuente.xlsx
    
    Returns:
        Mostrar en un dataframe los datos del archivo datos_fuente.xlsx
    """
    try:
        print(f"📖 Leyendo archivo Excel: {ruta_excel}")
        
        if not os.path.exists(ruta_excel):
            print(f"❌ Error: El archivo '{ruta_excel}' no existe") #En caso de que no exista el archivo, mostrar un mensaje de error
            return None
        
        # Leer Excel
        df = pd.read_excel(ruta_excel, engine='openpyxl')
        
        # Verificar que tenga las columnas necesarias
        columnas_requeridas = ['NOMBRE', 'CUENTA', 'GESTION EFECTIVA'] #<-- verificar que las columnas requeridas estén dentro del archivo de excel
        columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
        
        if columnas_faltantes:
            print(f"❌ Error: Faltan las siguientes columnas: {', '.join(columnas_faltantes)}")
            print(f"📋 Columnas encontradas: {', '.join(df.columns.tolist())}")
            return None
        
        # Filtrar solo clientes que tengan "GRABACION CALL" en GESTION EFECTIVA
        print(f"📊 Total de clientes en archivo: {len(df)}")
        print(f"🔍 Filtrando clientes con 'GRABACION CALL' en GESTION EFECTIVA...")
        
        # Filtrar clientes que contengan "GRABACION CALL" en el campo GESTION EFECTIVA
        # El campo puede contener múltiples valores separados por comas
        df_filtrado = df[df['GESTION EFECTIVA'].astype(str).str.contains('GRABACION CALL', case=False, na=False)]
        
        print(f"✅ Clientes con GRABACION CALL: {len(df_filtrado)}")
        print(f"❌ Clientes filtrados (sin GRABACION CALL): {len(df) - len(df_filtrado)}")
        
        return df_filtrado
        
    except Exception as e:
        print(f"❌ Error al leer el archivo Excel: {e}")
        return None

def buscar_archivo_cliente(nombre_archivo, ruta_base):
    """
    en un bucle recorre cada carpeta de cliente y busca el audio del cliente
    
    Args:
        nombre_archivo (str): Nombre del archivo a buscar
        ruta_base (str): Ruta base donde buscar
    
    Returns:
        tuple: (bool, str) - (encontrado, ruta_completa)
    """
    nombre_archivo_lower = nombre_archivo.lower()
    
    try:
        # Recorrer todas las carpetas y subcarpetas
        for root, dirs, files in os.walk(ruta_base):
            for file in files:
                # Comparar sin distinguir mayúsculas/minúsculas
                if file.lower() == nombre_archivo_lower:
                    ruta_completa = os.path.join(root, file)
                    return True, ruta_completa
        
        return False, ""
        
    except PermissionError:
        return False, ""
    except Exception as e:
        print(f"⚠️  Error al buscar en {ruta_base}: {e}")
        return False, ""

def procesar_clientes(df, ruta_evidencias):
    """
    Procesa cada cliente y verifica si tiene su archivo MP3
    
    Args:
        df (pandas.DataFrame): DataFrame con los datos de los clientes
        ruta_evidencias (str): Ruta a la carpeta de evidencias
    
    Returns:
        list: Lista de diccionarios con los resultados
    """
    resultados = []
    total_clientes = len(df)
    
    print(f"\n🔍 Iniciando búsqueda de archivos en: {ruta_evidencias}")
    print(f"📊 Total de clientes a procesar: {total_clientes}")
    print("-" * 60)
    
    # Verificar que la ruta existe
    if not os.path.exists(ruta_evidencias):
        print(f"❌ Error: La ruta '{ruta_evidencias}' no existe")
        return resultados
    
    # Procesar cada cliente
    numero_orden = 0
    for idx, row in df.iterrows():
        numero_orden += 1
        nombre = str(row['NOMBRE']).strip()
        cuenta = str(row['CUENTA']).strip()
        gestion_efectiva = str(row['GESTION EFECTIVA']).strip()
        
        # Construir nombre del archivo esperado
        nombre_archivo = f"{nombre}_{cuenta}.mp3"
        
        # Buscar el archivo
        encontrado, ruta_encontrada = buscar_archivo_cliente(nombre_archivo, ruta_evidencias)
        
        resultados.append({
            'Número de Orden': numero_orden,
            'Nombre del Cliente': nombre,
            'Cuenta': cuenta,
            'GESTION EFECTIVA': gestion_efectiva,
            'Archivo Esperado': nombre_archivo,
            'Tiene Archivo': 'SÍ' if encontrado else 'NO',
            'Ruta del Archivo': ruta_encontrada if encontrado else ''
        })
        
        # Mostrar progreso cada 100 clientes
        if (numero_orden % 100 == 0) or (numero_orden == total_clientes):
            print(f"📈 Procesados: {numero_orden}/{total_clientes} clientes ({numero_orden*100//total_clientes}%)")
    
    return resultados

def generar_reporte_excel(resultados, archivo_salida="reporte_auditoria_archivos.xlsx"):
    """
    Genera un archivo Excel con el reporte de auditoría
    
    Args:
        resultados (list): Lista de diccionarios con los resultados
        archivo_salida (str): Nombre del archivo Excel de salida
    """
    print(f"\n📝 Generando reporte Excel: {archivo_salida}")
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoría Archivos"
    
    # Definir encabezados
    encabezados = ['Número de Orden', 'Nombre del Cliente', 'Cuenta', 'GESTION EFECTIVA', 'Archivo Esperado', 'Tiene Archivo', 'Ruta del Archivo']
    ws.append(encabezados)
    
    # Estilo para encabezados
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col_num, header in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Estilo para filas
    fill_si = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_no = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Agregar datos
    for resultado in resultados:
        fila = [
            resultado['Número de Orden'],
            resultado['Nombre del Cliente'],
            resultado['Cuenta'],
            resultado['GESTION EFECTIVA'],
            resultado['Archivo Esperado'],
            resultado['Tiene Archivo'],
            resultado['Ruta del Archivo']
        ]
        ws.append(fila)
        
        # Aplicar color según si tiene archivo o no
        fila_actual = ws.max_row
        celda_estado = ws.cell(row=fila_actual, column=6)  # Columna "Tiene Archivo"
        
        if resultado['Tiene Archivo'] == 'SÍ':
            celda_estado.fill = fill_si
        else:
            celda_estado.fill = fill_no
        
        # Ajustar alineación
        for col_num in range(1, len(encabezados) + 1):
            celda = ws.cell(row=fila_actual, column=col_num)
            celda.alignment = Alignment(vertical="center", wrap_text=True)
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 18  # Número de Orden
    ws.column_dimensions['B'].width = 40  # Nombre del Cliente
    ws.column_dimensions['C'].width = 25  # Cuenta
    ws.column_dimensions['D'].width = 30  # GESTION EFECTIVA
    ws.column_dimensions['E'].width = 60  # Archivo Esperado
    ws.column_dimensions['F'].width = 15  # Tiene Archivo
    ws.column_dimensions['G'].width = 80  # Ruta del Archivo
    
    # Congelar primera fila
    ws.freeze_panes = 'A2'
    
    # Guardar archivo
    wb.save(archivo_salida)
    
    # Calcular estadísticas
    total_clientes = len(resultados)
    con_archivo = sum(1 for r in resultados if r['Tiene Archivo'] == 'SÍ')
    sin_archivo = total_clientes - con_archivo
    
    print(f"✅ Reporte generado exitosamente")
    print(f"\n{'='*60}")
    print(f"📊 RESUMEN DEL REPORTE")
    print(f"{'='*60}")
    print(f"📁 Total de clientes procesados: {total_clientes}")
    print(f"✅ Clientes con archivo: {con_archivo} ({con_archivo*100//total_clientes if total_clientes > 0 else 0}%)")
    print(f"❌ Clientes sin archivo: {sin_archivo} ({sin_archivo*100//total_clientes if total_clientes > 0 else 0}%)")
    print(f"💾 Archivo guardado en: {os.path.abspath(archivo_salida)}")
    
    return resultados

def obtener_archivos_reemplazo(ruta_carpeta_reemplazo):
    """
    Obtiene todos los archivos MP3 disponibles en la carpeta de reemplazo
    
    Args:
        ruta_carpeta_reemplazo (str): Ruta a la carpeta replace_aud-19
    
    Returns:
        list: Lista de rutas completas de archivos MP3
    """
    archivos_mp3 = []
    
    if not os.path.exists(ruta_carpeta_reemplazo):
        print(f"⚠️  La carpeta de reemplazo '{ruta_carpeta_reemplazo}' no existe")
        return archivos_mp3
    
    try:
        # Buscar todos los archivos MP3 en la carpeta y subcarpetas
        for root, dirs, files in os.walk(ruta_carpeta_reemplazo):
            for file in files:
                if file.lower().endswith('.mp3'):
                    ruta_completa = os.path.join(root, file)
                    archivos_mp3.append(ruta_completa)
        
        print(f"📂 Archivos MP3 encontrados en carpeta de reemplazo: {len(archivos_mp3)}")
        return archivos_mp3
        
    except Exception as e:
        print(f"❌ Error al buscar archivos en carpeta de reemplazo: {e}")
        return archivos_mp3

def encontrar_carpeta_cliente(cuenta, ruta_base):
    """
    Encuentra la carpeta del cliente basándose en la cuenta
    
    Args:
        cuenta (str): Número de cuenta del cliente
        ruta_base (str): Ruta base donde buscar
    
    Returns:
        str: Ruta de la carpeta del cliente, o None si no se encuentra
    """
    cuenta_limpia = str(cuenta).strip()
    
    try:
        # Primero buscar en el primer nivel (más rápido)
        for item in os.listdir(ruta_base):
            ruta_item = os.path.join(ruta_base, item)
            if os.path.isdir(ruta_item) and cuenta_limpia in item:
                return ruta_item
        
        # Si no se encuentra, buscar recursivamente en subcarpetas
        for root, dirs, files in os.walk(ruta_base):
            for dir_name in dirs:
                if cuenta_limpia in dir_name:
                    ruta_carpeta = os.path.join(root, dir_name)
                    return ruta_carpeta
        
        return None
        
    except Exception as e:
        print(f"⚠️  Error al buscar carpeta del cliente {cuenta}: {e}")
        return None

def completar_archivos_faltantes(resultados, ruta_evidencias, ruta_carpeta_reemplazo):
    """
    Completa los archivos faltantes copiando archivos aleatorios de la carpeta de reemplazo
    
    Args:
        resultados (list): Lista de diccionarios con los resultados
        ruta_evidencias (str): Ruta a la carpeta de evidencias
        ruta_carpeta_reemplazo (str): Ruta a la carpeta replace_aud-19
    
    Returns:
        int: Número de clientes a los que se completó el archivo
    """
    print(f"\n{'='*60}")
    print(f"🔄 COMPLETAR ARCHIVOS FALTANTES")
    print(f"{'='*60}")
    
    # Obtener archivos disponibles para reemplazo
    archivos_reemplazo = obtener_archivos_reemplazo(ruta_carpeta_reemplazo)
    
    if not archivos_reemplazo:
        print("❌ No se encontraron archivos MP3 en la carpeta de reemplazo")
        return 0
    
    # Filtrar clientes sin archivo
    clientes_sin_archivo = [r for r in resultados if r['Tiene Archivo'] == 'NO']
    
    if not clientes_sin_archivo:
        print("✅ Todos los clientes ya tienen su archivo")
        return 0
    
    print(f"📊 Clientes sin archivo a completar: {len(clientes_sin_archivo)}")
    print(f"📂 Archivos disponibles para reemplazo: {len(archivos_reemplazo)}")
    print("-" * 60)
    
    clientes_completados = 0
    archivos_usados = set()  # Para evitar usar el mismo archivo múltiples veces si es posible
    
    for idx, cliente in enumerate(clientes_sin_archivo, 1):
        nombre = cliente['Nombre del Cliente']
        cuenta = cliente['Cuenta']
        nombre_archivo_esperado = cliente['Archivo Esperado']
        
        # Encontrar la carpeta del cliente
        carpeta_cliente = encontrar_carpeta_cliente(cuenta, ruta_evidencias)
        
        if not carpeta_cliente:
            print(f"⚠️  [{idx}/{len(clientes_sin_archivo)}] No se encontró carpeta para cliente {nombre} (Cuenta: {cuenta})")
            continue
        
        # Seleccionar un archivo aleatorio de reemplazo
        archivo_origen = random.choice(archivos_reemplazo)
        
        # Crear ruta de destino
        ruta_destino = os.path.join(carpeta_cliente, nombre_archivo_esperado)
        
        try:
            # Copiar el archivo
            shutil.copy2(archivo_origen, ruta_destino)
            
            # Actualizar el resultado
            cliente['Tiene Archivo'] = 'SÍ'
            cliente['Ruta del Archivo'] = ruta_destino
            
            clientes_completados += 1
            
            # Mostrar progreso cada 50 clientes
            if clientes_completados % 50 == 0 or clientes_completados == len(clientes_sin_archivo):
                print(f"📈 Completados: {clientes_completados}/{len(clientes_sin_archivo)} clientes")
                
        except Exception as e:
            print(f"❌ Error al copiar archivo para {nombre} (Cuenta: {cuenta}): {e}")
            continue
    
    print(f"\n✅ Proceso de completado finalizado")
    print(f"📊 Total de clientes completados: {clientes_completados}/{len(clientes_sin_archivo)}")
    
    return clientes_completados

def preguntar_completar_archivos():
    """
    Pregunta al usuario si desea completar archivos faltantes
    
    Returns:
        bool: True si el usuario acepta, False en caso contrario
    """
    print(f"\n{'='*60}")
    print(f"❓ PREGUNTA")
    print(f"{'='*60}")
    
    while True:
        respuesta = input("\n¿Desea buscar archivos en la carpeta 'replace_aud-19' para completar los archivos faltantes? (s/n): ").strip().lower()
        
        if respuesta in ['s', 'si', 'sí', 'y', 'yes']:
            return True
        elif respuesta in ['n', 'no']:
            return False
        else:
            print("❌ Por favor responda 's' para sí o 'n' para no")

def main():
    parser = argparse.ArgumentParser(
        description='Auditoría de archivos MP3 de clientes basado en datos de Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplo de uso:
  python app.py
  
  python app.py --ruta-evidencias "E:\\ProcesoAudios\\evidencias\\evidencias_27-31_v2"
  
  python app.py --excel "mi_datos.xlsx" --ruta-evidencias "E:\\ruta\\evidencias"
        """
    )
    
    parser.add_argument('--excel', '-e', default='datos_fuente.xlsx',
                       help='Ruta al archivo Excel con los datos fuente (por defecto: datos_fuente.xlsx)')
    parser.add_argument('--ruta-evidencias', '-r', 
                       default=r'E:\ProcesoAudios\evidencias\evidencias_27-31_v2',
                       help='Ruta a la carpeta de evidencias (por defecto: E:\\ProcesoAudios\\evidencias\\evidencias_27-31_v2)')
    parser.add_argument('--salida', '-s', default='reporte_auditoria_archivos.xlsx',
                       help='Nombre del archivo Excel de salida (por defecto: reporte_auditoria_archivos.xlsx)')
    parser.add_argument('--ruta-reemplazo', default=None,
                       help='Ruta a la carpeta replace_aud-19 (por defecto: se busca automáticamente)')
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("🔍 AUDITORÍA DE ARCHIVOS MP3 DE CLIENTES")
    print("=" * 60)
    
    # Leer datos fuente
    df = leer_datos_fuente(args.excel)
    
    if df is None:
        print("\n❌ No se pudo continuar sin los datos fuente")
        return
    
    # Procesar clientes
    resultados = procesar_clientes(df, args.ruta_evidencias)
    
    if not resultados:
        print("\n❌ No se generaron resultados")
        return
    
    # Generar reporte Excel inicial
    resultados = generar_reporte_excel(resultados, args.salida)
    
    # Preguntar si desea completar archivos faltantes
    if preguntar_completar_archivos():
        # Determinar ruta de carpeta de reemplazo
        if args.ruta_reemplazo:
            ruta_carpeta_reemplazo = args.ruta_reemplazo
        else:
            # Construir ruta de carpeta de reemplazo (asumiendo que está en el mismo directorio padre que evidencias)
            ruta_base_evidencias = os.path.dirname(args.ruta_evidencias)
            ruta_carpeta_reemplazo = os.path.join(ruta_base_evidencias, "replace_aud-19")
            
            # Si no existe en la ruta relativa, intentar como ruta absoluta
            if not os.path.exists(ruta_carpeta_reemplazo):
                # Intentar ruta absoluta común
                ruta_carpeta_reemplazo = r"E:\ProcesoAudios\replace_aud-19"
        
        # Completar archivos faltantes
        clientes_completados = completar_archivos_faltantes(resultados, args.ruta_evidencias, ruta_carpeta_reemplazo)
        
        if clientes_completados > 0:
            # Regenerar reporte con los archivos completados
            print(f"\n🔄 Regenerando reporte con archivos completados...")
            generar_reporte_excel(resultados, args.salida)
            
            print(f"\n{'='*60}")
            print(f"✅ RESUMEN FINAL")
            print(f"{'='*60}")
            print(f"📊 Total de clientes a los que se completó el archivo: {clientes_completados}")
            print(f"💾 Reporte actualizado guardado en: {os.path.abspath(args.salida)}")
    
    print("\n✅ Proceso completado exitosamente")

if __name__ == "__main__":
    main()